import tkinter as tk
from tkinter import messagebox, ttk, Menu
from ttkthemes import ThemedStyle
from tkcalendar import DateEntry
import openpyxl
import Category


def is_float(value):
    try:
        float(value)
        return True
    except ValueError:
        return False


def change_theme(theme_name):
    style.set_theme(theme_name)
    root.update()


def on_theme_select():
    selected_theme = theme_var.get()
    change_theme(selected_theme)


def save_form(name_label, name_input,
              category_label, category_input,
              date_label,
              date_input, desc_label, desc_input,
              price_label, price_input):
    name_value = name_input.get()
    category_value = category_input.get()
    date_value = date_input.get()
    desc_value = desc_input.get("1.0", "end-1c")
    price_value = price_input.get()
    selected_option = category_input.get()

    # Sprawdź, czy pola nie są puste
    if not name_value or not date_value or not price_value:
        tk.messagebox.showerror("Błąd", "Wszystkie pola muszą być wypełnione.")
        return

    # Sprawdź, czy cena jest liczbą zmiennoprzecinkową
    if not is_float(price_value):
        tk.messagebox.showerror("Błąd", "Cena musi być liczbą zmiennoprzecinkową.")
        return
    if not selected_option:
        tk.messagebox.showerror("Błąd", "Wybierz kategorię z listy rozwijalnej.")
        return False

    form_data = {
        name_label.cget("text"): name_value,
        category_label.cget("text"): category_value,
        date_label.cget("text"): date_value,
        desc_label.cget("text"): desc_value,
        price_label.cget("text"): price_value
    }

    try:
        wb = openpyxl.load_workbook("dane.xlsx")
        ws = wb.active
    except FileNotFoundError:
        wb = openpyxl.Workbook()
        ws = wb.active

        headers = list(form_data.keys())
        ws.append(headers)

    data_values = list(form_data.values())
    ws.append(data_values)
    wb.save("dane.xlsx")
    messagebox.showinfo("Sukces", "Dane zapisane")


def show_form():
    form_window = tk.Toplevel(root)
    form_window.title("Formularz")
    form_window.geometry("400x800")
    name_label = ttk.Label(form_window, text="NAZWA")
    categories = ["Aldi", "Rossman", "Netto", "Żabka", "Dealz", "E-dym", "Liquder", "Costa", "Starbucks", "Inne"]
    category_label = ttk.Label(form_window, text="Kategoria")

    category_input = ttk.Combobox(form_window, values=categories)

    name_input = ttk.Entry(form_window)
    date_label = ttk.Label(form_window, text="DATA")
    date_input = DateEntry(form_window, background=style, bd=2, date_pattern='yyyy.mm.dd')
    desc_label = ttk.Label(form_window, text="OPIS")
    desc_input = tk.Text(form_window)
    price_label = ttk.Label(form_window, text='SUMA')
    price_input = ttk.Entry(form_window)

    name_label.pack(fill=tk.BOTH, pady=10)
    name_input.pack(fill=tk.BOTH, padx=10)

    category_label.pack(fill=tk.BOTH, padx=10)
    category_input.pack(fill=tk.BOTH, padx=10)

    date_label.pack(fill=tk.BOTH, pady=10)
    date_input.pack(fill=tk.BOTH, padx=10)

    desc_label.pack(fill=tk.BOTH, padx=10)
    desc_input.pack(fill=tk.BOTH, padx=10)

    price_label.pack(fill=tk.BOTH, padx=10)
    price_input.pack(fill=tk.BOTH, padx=10)

    save_button = ttk.Button(form_window, text="Zapisz", command=lambda: save_form(name_label, name_input,
                                                                                   category_label, category_input,
                                                                                   date_label,
                                                                                   date_input, desc_label, desc_input,
                                                                                   price_label, price_input))
    save_button.pack(fill=tk.BOTH, padx=12, pady=5)


def exit_app():
    root.destroy()


root = tk.Tk()
style = ThemedStyle(root)
theme_var = tk.StringVar()

style.set_theme("radiance")
available_themes = style.theme_names()

menu = tk.Frame(root)
menu.pack(pady=10)

root.title("Moje rachunki")
root.geometry("700x100")


def show_dashboard():
    dashboard_window = tk.Toplevel(root)
    dashboard_window.geometry("800x600")
    update_categories_summary()
    total_price_label = tk.Label(dashboard_window, text="Suma:")
    total_price_label.pack()

    try:
        wb = openpyxl.load_workbook('dane.xlsx')
        ws = wb.active
        prices = [float(cell.value) for cell in ws['E'][1:] if cell.value]
        total_price = sum(prices)
        total_price_label.config(text=f"Suma wydatków: {total_price} zł ")
    except FileNotFoundError:
        tk.messagebox.showerror("Błąd", "Plik dane.xlsx nie istnieje.")

    except Exception as e:
        tk.messagebox.showerror("Błąd", f"Wystąpił błąd: {str(e)}")

    category_frame = tk.Frame(dashboard_window)
    category_frame.pack()
    try:
        wb = openpyxl.load_workbook('dane.xlsx')
        ws = wb.active
        category_sums = {}
        for row in ws.iter_rows(min_row=2, values_only=True):
            category, price = row[1], row[4]
            if category not in category_sums:
                category_sums[category] = 0
            category_sums[category] += float(price)
            for child in category_frame.winfo_children():
                child.destroy()

            for category, total_price in category_sums.items():
                category_label = tk.Label(category_frame, text=f"{category} {total_price}")
                category_label.pack()

    except FileNotFoundError:
        tk.messagebox.showerror("Błąd", "Plik dane.xlsx nie istnieje.")
    except Exception as e:
        tk.messagebox.showerror("Błąd", f"Wystąpił błąd: {str(e)}")


def update_categories_summary():
    pass


def add_category():
    category = Category
    category.add()
    pass


menubar = Menu(root)
menubar.add_command(label="Dashboard", command=lambda: show_dashboard())
menubar.add_command(label="Dodaj rachunek", command=lambda: show_form())

menubar.add_command(label="Dodaj nową kategorię", command=lambda: add_category())
menubar.add_command(label="Zamknij aplikację", command=lambda: exit_app())
hello_label = tk.Label(root, text="Witaj w aplikacji Moje rachunki")
hello_label.pack(fill=tk.BOTH, padx=10, pady=10)

root.config(menu=menubar)
root.mainloop()
