import tkinter as tk
from tkinter import ttk, messagebox
from ttkthemes import ThemedStyle
import openpyxl


def add():
    add_category_window = tk.Tk()
    add_category_window.title("Dodaj nową kategorie")
    add_category_window.geometry("500x200")
    style = ThemedStyle(add_category_window)

    style.set_theme("radiance")

    name_label = ttk.Label(add_category_window, text="Nazwa kategorii")
    name_label.pack(fill=tk.BOTH, padx=10, pady=5)
    name_input = ttk.Entry(add_category_window)
    name_input.pack(fill=tk.BOTH, padx=10, pady=5)

    save_button = ttk.Button(add_category_window, text="Dodaj", command=lambda: save_form(name_input))
    save_button.pack(fill=tk.BOTH, padx=10, pady=10)
    add_category_window.mainloop()
    pass


def save_form(name_input):
    name_value = name_input.get()
    if not name_value:
        tk.messagebox.showerror("Błąd", "Wszystkie pola muszą być wypełnione.")
        return
    try:
        wb = openpyxl.load_workbook('dane.xlsx')
    except FileNotFoundError:
        wb = openpyxl.Workbook()
    if "kategorie" not in wb.sheetnames:
        wb.create_sheet("kategorie")

    ws = wb['kategorie']
    header = "Kategorie"
    ws.append([header])

    ws.append([name_value])
    wb.save('dane.xlsx')

    messagebox.showinfo("Sukces", "Dane zapisane")

    pass


class Category:
    pass
